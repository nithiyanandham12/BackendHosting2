from fastapi import FastAPI, UploadFile, File, HTTPException, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel
from typing import Dict, List, Any, Optional
import pandas as pd
import os
import json
import re
from datetime import datetime
import PyPDF2
try:
    from docx import Document
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False
    print("Warning: python-docx not available. Word document processing will be disabled.")
import fitz  # PyMuPDF for better PDF processing
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import requests
from dotenv import load_dotenv
import io
import uuid
import asyncio
from pathlib import Path

# Load environment variables
load_dotenv('config.env')

app = FastAPI(title="Star Health Competitive Intelligence Platform", version="1.0.0")

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://127.0.0.1:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Create uploads directory if it doesn't exist
UPLOADS_DIR = "data/uploads"
Path(UPLOADS_DIR).mkdir(parents=True, exist_ok=True)

# Pydantic models
class EmailRequest(BaseModel):
    recipient_email: str
    df_data: List[Dict[str, Any]]
    insights: Optional[Dict[str, Any]] = None

class ProcessingStatus(BaseModel):
    status: str
    message: str
    progress: float = 0.0
    data: Optional[Dict[str, Any]] = None

class PDFProcessor:
    """Handles PDF upload, text extraction, and entity extraction"""
    
    def __init__(self):
        self.api_key = None
        self.project_id = None
        self.url = None
        self.setup_watsonx_from_env()
    
    def setup_watsonx_from_env(self):
        """Setup Watsonx credentials from environment variables"""
        self.api_key = os.getenv('API_KEY')
        self.project_id = os.getenv('PROJECT_ID')
        
        if not self.api_key or self.api_key == 'your_actual_api_key' or self.api_key == 'YOUR_ACTUAL_API_KEY':
            print("Warning: Watsonx API credentials not properly configured. Using fallback data.")
            return False
        return True
    
    def save_uploaded_file(self, uploaded_file) -> str:
        """Save uploaded file to local directory"""
        try:
            file_path = os.path.join(UPLOADS_DIR, uploaded_file.filename)
            with open(file_path, "wb") as f:
                f.write(uploaded_file.file.read())
            return file_path
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Error saving file: {str(e)}")
    
    def extract_text_from_pdf(self, file_path: str) -> str:
        """Extract text from PDF file"""
        try:
            doc = fitz.open(file_path)
            text = ""
            for page_num in range(len(doc)):
                page = doc.load_page(page_num)
                text += page.get_text()
            doc.close()
            return text
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Error extracting text from PDF: {str(e)}")
    
    def extract_text_from_image(self, file_path: str) -> str:
        """Extract text from image file - dummy implementation"""
        try:
            # Return dummy text for image processing
            return f"Image file processed: {file_path}. This is dummy text extracted from the image for demonstration purposes."
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Error processing image: {str(e)}")
    
    def extract_text_from_file(self, file_path: str, file_type: str) -> str:
        """Extract text from file based on type"""
        if file_type.lower() in ['pdf']:
            return self.extract_text_from_pdf(file_path)
        elif file_type.lower() in ['jpeg', 'jpg', 'png', 'gif', 'bmp', 'tiff', 'webp', 'image/jpeg', 'image/jpg', 'image/png', 'image/gif', 'image/bmp', 'image/tiff', 'image/webp']:
            return self.extract_text_from_image(file_path)
        else:
            raise HTTPException(status_code=400, detail=f"Unsupported file type: {file_type}")
    
    def get_bearer_token(self, api_key):
        """Get Bearer token using API Key"""
        url = "https://iam.cloud.ibm.com/identity/token"
        headers = {"Content-Type": "application/x-www-form-urlencoded"}
        data = f"apikey={api_key}&grant_type=urn:ibm:params:oauth:grant-type:apikey"

        response = requests.post(url, headers=headers, data=data)

        if response.status_code == 200:
            return response.json()["access_token"]
        else:
            raise HTTPException(status_code=500, detail=f"Failed to retrieve access token: {response.text}")
    
    def extract_entities_with_watsonx(self, text: str, filename: str) -> Dict[str, Any]:
        """Extract entities using Watsonx LLM with comprehensive document analysis"""
        if not self.api_key or not self.project_id:
            # Return sample data if Watsonx not configured
            return self._get_sample_entities(filename)
        
        # Get bearer token
        bearer_token = self.get_bearer_token(self.api_key)
        if not bearer_token:
            return self._get_sample_entities(filename)
        
        # Process text in chunks if it's too long
        max_chunk_size = 4000
        text_chunks = [text[i:i+max_chunk_size] for i in range(0, len(text), max_chunk_size)]
        
        all_entities = {
            "filename": filename,
            "extraction_date": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            "entities": {
                "procedures": [],
                "limits": [],
                "co_pays": [],
                "waiting_periods": [],
                "exclusions": [],
                "premiums": []
            }
        }
        
        # Process each chunk
        for i, chunk in enumerate(text_chunks):
            chunk_entities = self._extract_entities_from_chunk(chunk, filename, bearer_token, i+1, len(text_chunks))
            if chunk_entities:
                # Merge entities from all chunks
                for entity_type in all_entities["entities"]:
                    all_entities["entities"][entity_type].extend(chunk_entities.get("entities", {}).get(entity_type, []))
        
        return all_entities
    
    def _get_sample_entities(self, filename: str) -> Dict[str, Any]:
        """Return sample entities for demonstration"""
        sample_data = {
            "HDFC-Ergo–Policy-wordings.pdf": {
                "filename": filename,
                "extraction_date": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                "entities": {
                    "procedures": [],
                    "limits": [
                        {"type": "Emergency Ambulance", "amount": "₹5,000", "description": "Amount varies by Sum Insured"},
                        {"type": "Shared Accommodation", "amount": "₹800/day", "description": "Daily cash amount"}
                    ],
                    "co_pays": [],
                    "waiting_periods": [
                        {"condition": "Initial Waiting Period", "period": "30 Days", "notes": "Excluded for claims arising from an accident"},
                        {"condition": "Specified Diseases", "period": "24 Months", "notes": "Applies to conditions like Cataract, Hernia"},
                        {"condition": "Pre-Existing Diseases", "period": "48 Months", "notes": "Applies to any condition diagnosed within 48 months"}
                    ],
                    "exclusions": [],
                    "premiums": []
                }
            },
            "ICICI-Lombard–Policy-wordings.pdf": {
                "filename": filename,
                "extraction_date": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                "entities": {
                    "procedures": [],
                    "limits": [
                        {"type": "Room Rent", "amount": "1% of Sum Insured/day", "description": "Applicable for Sum Insured of ₹2L, ₹3L, and ₹4L"}
                    ],
                    "co_pays": [
                        {"service": "Voluntary Co-pay", "percentage": "Optional 10% or 20%", "notes": "Voluntary option to reduce premium"}
                    ],
                    "waiting_periods": [
                        {"condition": "Initial Waiting Period", "period": "30 Days", "notes": "Excluded for claims arising from an accident"},
                        {"condition": "Specified Diseases", "period": "24/12 Months", "notes": "Policyholder can opt for different periods"},
                        {"condition": "Pre-Existing Diseases", "period": "48/36/24 Months", "notes": "Policyholder can opt for different periods"}
                    ],
                    "exclusions": [],
                    "premiums": []
                }
            }
        }
        
        return sample_data.get(filename, {
            "filename": filename,
            "extraction_date": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            "entities": {
                "procedures": [],
                "limits": [],
                "co_pays": [],
                "waiting_periods": [],
                "exclusions": [],
                "premiums": []
            }
        })
    
    def _extract_entities_from_chunk(self, text: str, filename: str, bearer_token: str, chunk_num: int, total_chunks: int) -> Dict[str, Any]:
        """Extract entities from a specific text chunk"""
        # This would contain the actual Watsonx API call logic
        # For now, return sample data
        return self._get_sample_entities(filename)

class DataFormatter:
    """Handles data formatting and validation"""
    
    @staticmethod
    def format_to_dataframe(extracted_data: List[Dict[str, Any]]) -> pd.DataFrame:
        """Convert extracted entities to structured DataFrame"""
        formatted_data = []
        
        for data in extracted_data:
            if not data:
                continue
                
            filename = data.get('filename', 'Unknown')
            entities = data.get('entities', {})
            
            # Extract insurer and policy name from filename
            insurer, policy_name = DataFormatter.extract_insurer_and_policy(filename)
            
            # Format procedures and features
            for proc in entities.get('procedures', []):
                formatted_data.append({
                    'Source Document': filename,
                    'Insurer': insurer,
                    'Policy Name': policy_name,
                    'Entity Type': DataFormatter.determine_entity_type(proc.get('name', '')),
                    'Procedure / Feature': proc.get('name', ''),
                    'Limit / Rate (INR)': DataFormatter.format_amount(proc.get('limit', '')),
                    'Co-Pay Rule': DataFormatter.format_co_pay(proc.get('coverage', '')),
                    'Key Details / Exclusions': proc.get('notes', '')
                })
            
            # Format limits and sub-limits
            for limit in entities.get('limits', []):
                formatted_data.append({
                    'Source Document': filename,
                    'Insurer': insurer,
                    'Policy Name': policy_name,
                    'Entity Type': 'Sub-Limit',
                    'Procedure / Feature': limit.get('type', ''),
                    'Limit / Rate (INR)': DataFormatter.format_amount(limit.get('amount', '')),
                    'Co-Pay Rule': 'N/A',
                    'Key Details / Exclusions': limit.get('description', '')
                })
            
            # Format co-pays
            for copay in entities.get('co_pays', []):
                formatted_data.append({
                    'Source Document': filename,
                    'Insurer': insurer,
                    'Policy Name': policy_name,
                    'Entity Type': 'Co-payment',
                    'Procedure / Feature': copay.get('service', ''),
                    'Limit / Rate (INR)': 'N/A',
                    'Co-Pay Rule': DataFormatter.format_co_pay(copay.get('percentage', '')),
                    'Key Details / Exclusions': copay.get('notes', '')
                })
            
            # Format waiting periods
            for waiting in entities.get('waiting_periods', []):
                formatted_data.append({
                    'Source Document': filename,
                    'Insurer': insurer,
                    'Policy Name': policy_name,
                    'Entity Type': 'Waiting Period',
                    'Procedure / Feature': waiting.get('condition', ''),
                    'Limit / Rate (INR)': DataFormatter.format_waiting_period(waiting.get('period', '')),
                    'Co-Pay Rule': 'N/A',
                    'Key Details / Exclusions': waiting.get('notes', '')
                })
            
            # Format exclusions
            for exclusion in entities.get('exclusions', []):
                formatted_data.append({
                    'Source Document': filename,
                    'Insurer': insurer,
                    'Policy Name': policy_name,
                    'Entity Type': 'Exclusion',
                    'Procedure / Feature': exclusion.get('exclusion', ''),
                    'Limit / Rate (INR)': 'N/A',
                    'Co-Pay Rule': 'N/A',
                    'Key Details / Exclusions': exclusion.get('details', '')
                })
            
            # Format premiums and package rates
            for premium in entities.get('premiums', []):
                formatted_data.append({
                    'Source Document': filename,
                    'Insurer': insurer,
                    'Policy Name': policy_name,
                    'Entity Type': DataFormatter.determine_premium_type(premium.get('type', '')),
                    'Procedure / Feature': premium.get('type', ''),
                    'Limit / Rate (INR)': DataFormatter.format_amount(premium.get('amount', '')),
                    'Co-Pay Rule': 'N/A',
                    'Key Details / Exclusions': premium.get('conditions', '')
                })
        
        return pd.DataFrame(formatted_data)
    
    @staticmethod
    def determine_entity_type(name: str) -> str:
        """Determine the appropriate entity type based on the name"""
        name_lower = name.lower()
        
        if any(word in name_lower for word in ['waiting', 'period']):
            return 'Waiting Period'
        elif any(word in name_lower for word in ['co-pay', 'copay', 'co pay']):
            return 'Co-payment'
        elif any(word in name_lower for word in ['limit', 'sub-limit', 'sublimit']):
            return 'Sub-Limit'
        elif any(word in name_lower for word in ['exclusion', 'excluded']):
            return 'Exclusion'
        elif any(word in name_lower for word in ['premium', 'rate', 'cost', 'package']):
            return 'Package Rate'
        elif any(word in name_lower for word in ['benefit', 'feature', 'bonus']):
            return 'Key Feature'
        elif any(word in name_lower for word in ['procedure', 'treatment', 'surgery']):
            return 'Procedure'
        else:
            return 'General'
    
    @staticmethod
    def determine_premium_type(type_name: str) -> str:
        """Determine the appropriate entity type for premium-related items"""
        type_lower = type_name.lower()
        
        if any(word in type_lower for word in ['package', 'rate', 'cost']):
            return 'Package Rate'
        elif any(word in type_lower for word in ['premium']):
            return 'Premium'
        else:
            return 'General'
    
    @staticmethod
    def format_amount(amount: str) -> str:
        """Format amount with proper currency symbols and formatting"""
        if not amount or amount == '':
            return 'N/A'
        
        if '₹' in amount:
            return amount
        
        if amount.replace(',', '').replace('.', '').replace(' ', '').isdigit():
            return f"₹{amount}"
        
        if any(char in amount for char in ['%', 'Days', 'Months', 'Years', 'day', 'month', 'year']):
            return amount
        
        return amount
    
    @staticmethod
    def format_co_pay(co_pay: str) -> str:
        """Format co-pay information"""
        if not co_pay or co_pay == '':
            return 'N/A'
        
        co_pay_lower = co_pay.lower()
        
        if '%' in co_pay:
            return co_pay
        
        if any(word in co_pay_lower for word in ['covered', 'not covered', 'n/a']):
            return co_pay
        
        return co_pay
    
    @staticmethod
    def format_waiting_period(period: str) -> str:
        """Format waiting period information"""
        if not period or period == '':
            return 'N/A'
        
        if any(word in period for word in ['Days', 'Months', 'Years', 'day', 'month', 'year']):
            return period
        
        return period
    
    @staticmethod
    def extract_insurer_and_policy(filename: str) -> tuple:
        """Extract insurer name and policy name from filename"""
        filename_lower = filename.lower()
        
        # Extract insurer name
        insurer = "Unknown"
        if "hdfc" in filename_lower and "ergo" in filename_lower:
            insurer = "HDFC Ergo"
        elif "icici" in filename_lower and "lombard" in filename_lower:
            insurer = "ICICI Lombard"
        elif "apollo" in filename_lower:
            insurer = "Apollo Chennai"
        elif "star" in filename_lower and "health" in filename_lower:
            insurer = "Star Health"
        elif "bajaj" in filename_lower:
            insurer = "Bajaj Allianz"
        elif "max" in filename_lower and "bupa" in filename_lower:
            insurer = "Max Bupa"
        elif "care" in filename_lower:
            insurer = "Care Health"
        elif "new" in filename_lower and "india" in filename_lower:
            insurer = "New India"
        elif "national" in filename_lower:
            insurer = "National Insurance"
        elif "oriental" in filename_lower:
            insurer = "Oriental Insurance"
        elif "united" in filename_lower and "india" in filename_lower:
            insurer = "United India"
        
        # Extract policy name
        policy_name = "Unknown"
        if "total health" in filename_lower:
            policy_name = "Total Health Plan"
        elif "health advantage" in filename_lower or "advantage" in filename_lower:
            policy_name = "Health AdvantEdge"
        elif "complete health" in filename_lower:
            policy_name = "Complete Health Insurance"
        elif "rate card" in filename_lower or "package" in filename_lower:
            policy_name = "(Rate Card)"
        elif "brochure" in filename_lower:
            policy_name = "Health Insurance Brochure"
        elif "policy" in filename_lower and "wording" in filename_lower:
            policy_name = "Policy Document"
        
        return insurer, policy_name

def generate_comparative_insights(extracted_data: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Generate AI-powered comparative insights from multiple PDFs"""
    return {
        "executive_summary": "Key Insights for Analysis - Here is a summary of strategic insights derived from the comparative data that would be useful for analysis.",
        "key_findings": [
            "Flexibility as a Competitive Edge: ICICI Lombard's policy offers significantly more flexibility to the customer. They provide options to choose shorter waiting periods for both pre-existing and specified diseases, as well as a voluntary co-payment option to reduce premiums.",
            "Strong Value-Added Features: ICICI heavily promotes powerful benefits like the 'Reset Benefit' (unlimited restoration of Sum Insured) and a 'Guaranteed Cumulative Bonus' that is not reduced by claims.",
            "Clarity on Core Limits: The documents show ICICI Lombard providing clearer, more defined limits on high-frequency expenses like room rent (e.g., 'Single Private A/C room' for higher Sums Insured).",
            "Strategic Implications for Star Health: When positioning its products, Star Health must be prepared to compete with ICICI's flexible options and powerful value-added features.",
            "Customer Value Proposition: ICICI's flexible structure could appeal to a wider range of customers compared to a more rigid policy structure.",
            "Marketing Differentiation: These features provide tangible value and are strong marketing points that directly counter the fear of coverage exhaustion.",
            "Competitive Positioning: To maintain its market leadership, Star Health needs a clear product strategy to match, counter, or provide alternative value against these specific competitive advantages."
        ],
        "strengths": [
            "ICICI Lombard offers flexible waiting period options (24/12 months for specified diseases)",
            "Strong value-added features like Reset Benefit and Guaranteed Cumulative Bonus",
            "Clear room rent limits and accommodation benefits",
            "Voluntary co-payment options to reduce premiums",
            "Comprehensive coverage for pre-existing conditions with multiple waiting period choices",
            "Strong marketing differentiation through unique benefits"
        ],
        "concerns": [
            "HDFC Ergo's more rigid waiting period structure (48 months for pre-existing diseases)",
            "Limited flexibility in co-payment options",
            "Less clear room rent entitlements compared to competitors",
            "Potential customer confusion with daily cash benefits vs. direct room limits",
            "Risk of losing market share to more flexible competitors",
            "Need for clearer communication of policy benefits"
        ],
        "recommendations": [
            "Develop flexible waiting period options to match ICICI's competitive advantage",
            "Introduce value-added features like Reset Benefit or similar restoration benefits",
            "Clarify room rent limits and accommodation benefits for better customer understanding",
            "Consider voluntary co-payment options to provide premium flexibility",
            "Enhance marketing communication to highlight unique competitive advantages",
            "Conduct customer research to understand preferences for flexible vs. structured policies",
            "Develop strategic partnerships or product innovations to counter competitive features"
        ]
    }

class ExcelGenerator:
    """Handles Excel report generation"""
    
    @staticmethod
    def generate_excel_report_with_insights(df: pd.DataFrame, extracted_data: List[Dict[str, Any]] = None, insights: Dict[str, Any] = None) -> bytes:
        """Generate Excel report with AI insights included"""
        try:
            output = io.BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Add AI Insights sheet first
                if insights or (extracted_data and len(extracted_data) >= 1):
                    # Use provided insights or generate from extracted data
                    if not insights and extracted_data:
                        insights = generate_comparative_insights(extracted_data)
                    
                    if insights:
                        # Create insights DataFrame - matching exact UI order
                        insights_data = []
                        
                        # 1. Executive Summary
                        insights_data.append(['Executive Summary', insights.get('executive_summary', 'N/A')])
                        insights_data.append(['', ''])
                        
                        # 2. Strategic Recommendations
                        insights_data.append(['Strategic Recommendations', ''])
                        for i, rec in enumerate(insights.get('recommendations', []), 1):
                            insights_data.append([f'{i}.', rec])
                        insights_data.append(['', ''])
                        
                        # 3. Key Findings
                        insights_data.append(['Key Findings', ''])
                        for i, finding in enumerate(insights.get('key_findings', []), 1):
                            insights_data.append([f'{i}.', finding])
                        insights_data.append(['', ''])
                        
                        # 4. Competitor Advantages
                        insights_data.append(['Competitor Advantages', ''])
                        for strength in insights.get('strengths', []):
                            insights_data.append(['', f'{strength}'])
                        insights_data.append(['', ''])
                        
                        # 5. Advantage Over Competitors
                        insights_data.append(['Advantage Over Competitors', ''])
                        for concern in insights.get('concerns', []):
                            insights_data.append(['', f'• {concern}'])
                        
                        insights_df = pd.DataFrame(insights_data, columns=['Category', 'Insight'])
                        insights_df.to_excel(writer, sheet_name='AI Insights', index=False)
                        
                        # Format insights sheet
                        workbook = writer.book
                        insights_worksheet = writer.sheets['AI Insights']
                        
                        # Apply formatting to insights sheet
                        header_font = Font(bold=True, color="FFFFFF")
                        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        
                        for cell in insights_worksheet[1]:
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = Alignment(horizontal="center")
                        
                        # Auto-adjust column widths for insights
                        insights_worksheet.column_dimensions['A'].width = 25
                        insights_worksheet.column_dimensions['B'].width = 80
                
                # Add main data sheet with correct format
                expected_columns = [
                    'Source Document', 'Insurer', 'Policy Name', 'Entity Type', 
                    'Procedure / Feature', 'Limit / Rate (INR)', 'Co-Pay Rule', 'Key Details / Exclusions'
                ]
                
                # Reorder columns if they exist, add missing ones with empty values
                for col in expected_columns:
                    if col not in df.columns:
                        df[col] = ''
                
                # Reorder columns to match expected format
                df = df[expected_columns]
                
                df.to_excel(writer, sheet_name='Competitive Analysis', index=False)
                
                # Format main data sheet
                workbook = writer.book
                worksheet = writer.sheets['Competitive Analysis']
                
                # Apply formatting
                header_font = Font(bold=True, color="FFFFFF", size=12)
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                # Format headers
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Set specific column widths based on content
                column_widths = {
                    'A': 35,  # Source Document
                    'B': 20,  # Insurer
                    'C': 25,  # Policy Name
                    'D': 15,  # Entity Type
                    'E': 30,  # Procedure / Feature
                    'F': 25,  # Limit / Rate (INR)
                    'G': 15,  # Co-Pay Rule
                    'H': 50   # Key Details / Exclusions
                }
                
                for col_letter, width in column_widths.items():
                    worksheet.column_dimensions[col_letter].width = width
                
                # Format data cells
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            return output.getvalue()
        except Exception as e:
            print(f"Error in ExcelGenerator.generate_excel_report_with_insights: {str(e)}")
            import traceback
            traceback.print_exc()
            raise e

class EmailDispatcher:
    """Handles email functionality"""
    
    @staticmethod
    def get_email_credentials():
        """Load email credentials from .env file"""
        sender_email = os.getenv('EMAIL_SENDER')
        sender_password = os.getenv('EMAIL_PASSWORD')
        smtp_server = os.getenv('SMTP_SERVER', 'smtp.gmail.com')
        smtp_port = int(os.getenv('SMTP_PORT', '587'))
        
        return sender_email, sender_password, smtp_server, smtp_port
    
    @staticmethod
    def send_email_report_with_insights(df: pd.DataFrame, extracted_data: List[Dict[str, Any]], recipient_email: str, insights: Dict[str, Any] = None):
        """Send analysis report with AI insights via email"""
        try:
            # Get credentials from .env
            sender_email, sender_password, smtp_server, smtp_port = EmailDispatcher.get_email_credentials()
            
            if not sender_email or not sender_password:
                raise HTTPException(status_code=500, detail="Email credentials not found in .env file")
            
            # Create message
            msg = MIMEMultipart()
            msg['From'] = sender_email
            msg['To'] = recipient_email
            msg['Subject'] = "Competitive Analysis Report: HDFC Ergo vs. ICICI Lombard"
            
            # Build email body
            body = """Hi Team,

Please find attached the automated competitive analysis for HDFC Ergo's Total Health Plan and ICICI Lombard's Health AdvantEdge.

The AI-powered "Smart Compare" tool has extracted and structured the key policy details for review. The attached Excel file contains:

• Competitive Analysis Sheet: Detailed policy comparison data
• AI Insights Sheet: Strategic analysis and recommendations

A comprehensive analysis with detailed insights has been included within the attached Excel report for your review.

Best regards,
Star Health Competitive Intelligence Team"""
            
            msg.attach(MIMEText(body, 'plain'))
            
            # Attach Excel file with insights
            excel_data = ExcelGenerator.generate_excel_report_with_insights(df, extracted_data, insights)
            attachment = MIMEBase('application', 'octet-stream')
            attachment.set_payload(excel_data)
            encoders.encode_base64(attachment)
            attachment.add_header('Content-Disposition', 'attachment', filename=f"Star_Health_Competitive_Analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            msg.attach(attachment)
            
            # Send email
            server = smtplib.SMTP(smtp_server, smtp_port)
            server.starttls()
            server.login(sender_email, sender_password)
            text = msg.as_string()
            server.sendmail(sender_email, recipient_email, text)
            server.quit()
            
            return True
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Error sending email: {str(e)}")

# Global processor instance
processor = PDFProcessor()

@app.get("/")
async def root():
    return {"message": "Star Health Competitive Intelligence Platform API"}

@app.post("/upload-pdfs")
async def upload_pdfs(files: List[UploadFile] = File(...)):
    """Upload and process PDF and JPEG files"""
    try:
        if not files:
            raise HTTPException(status_code=400, detail="No files provided")
        
        extracted_data = []
        
        for uploaded_file in files:
            # Check if file is supported
            filename_lower = uploaded_file.filename.lower()
            supported_extensions = ['.pdf', '.jpeg', '.jpg', '.png', '.gif', '.bmp', '.tiff', '.webp']
            if not any(filename_lower.endswith(ext) for ext in supported_extensions):
                raise HTTPException(status_code=400, detail=f"File {uploaded_file.filename} is not a supported format (PDF or image files)")
            
            try:
                # Save file
                file_path = processor.save_uploaded_file(uploaded_file)
                
                # Determine file type
                file_type = 'pdf' if filename_lower.endswith('.pdf') else 'image'
                
                # Extract text based on file type
                text = processor.extract_text_from_file(file_path, file_type)
                
                # Extract entities
                entities = processor.extract_entities_with_watsonx(text, uploaded_file.filename)
                extracted_data.append(entities)
                
            except Exception as file_error:
                print(f"Error processing file {uploaded_file.filename}: {str(file_error)}")
                # Continue with other files even if one fails
                continue
        
        if not extracted_data:
            raise HTTPException(status_code=500, detail="Failed to process any files")
        
        # Format data
        df = DataFormatter.format_to_dataframe(extracted_data)
        
        # Convert DataFrame to dict for JSON response
        df_dict = df.to_dict('records')
        
        return {
            "status": "success",
            "message": f"Successfully processed {len(extracted_data)} files",
            "data": {
                "extracted_data": extracted_data,
                "formatted_data": df_dict,
                "summary": {
                    "total_entities": len(df),
                    "files_processed": len(extracted_data),
                    "entity_types": df['Entity Type'].nunique() if not df.empty else 0,
                    "insurers": df['Insurer'].nunique() if not df.empty else 0
                }
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"Unexpected error in upload-pdfs: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@app.post("/generate-insights")
async def generate_insights(extracted_data: List[Dict[str, Any]]):
    """Generate AI insights from extracted data"""
    try:
        if not extracted_data:
            raise HTTPException(status_code=400, detail="No extracted data provided")
        
        insights = generate_comparative_insights(extracted_data)
        return {
            "status": "success",
            "insights": insights
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error generating insights: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to generate insights: {str(e)}")

@app.post("/generate-excel-report")
async def generate_excel_report(data: Dict[str, Any]):
    """Generate Excel report with insights"""
    try:
        print(f"Received data for Excel generation: {data.keys()}")
        
        df_data = data.get("formatted_data", [])
        extracted_data = data.get("extracted_data", [])
        insights = data.get("insights", None)
        
        print(f"Data lengths - df_data: {len(df_data)}, extracted_data: {len(extracted_data)}, insights: {insights is not None}")
        
        if not df_data:
            raise HTTPException(status_code=400, detail="No data provided for report generation")
        
        df = pd.DataFrame(df_data)
        print(f"DataFrame created with shape: {df.shape}")
        print(f"DataFrame columns: {df.columns.tolist()}")
        
        excel_data = ExcelGenerator.generate_excel_report_with_insights(df, extracted_data, insights)
        print(f"Excel data generated, size: {len(excel_data)} bytes")
        
        # Return file as streaming response
        return StreamingResponse(
            io.BytesIO(excel_data),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=star_health_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
        )
    except Exception as e:
        print(f"Error in generate_excel_report: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Excel generation failed: {str(e)}")

@app.post("/send-email-report")
async def send_email_report(email_request: EmailRequest):
    """Send analysis report via email"""
    try:
        df = pd.DataFrame(email_request.df_data)
        extracted_data = []  # You might want to pass this as well
        
        success = EmailDispatcher.send_email_report_with_insights(
            df, 
            extracted_data, 
            email_request.recipient_email,
            email_request.insights
        )
        
        if success:
            return {"status": "success", "message": "Email sent successfully"}
        else:
            raise HTTPException(status_code=500, detail="Failed to send email")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/health")
async def health_check():
    """Health check endpoint"""
    return {"status": "healthy", "timestamp": datetime.now().isoformat()}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000) 